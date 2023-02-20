import openpyxl

wb = openpyxl.load_workbook('store.xlsx')


print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

sheet = wb['Products']


sheet = wb.active


b2_cell = sheet['B2']
c2_cell = sheet['c2']


print(b2_cell.value, c2_cell.value)


print(sheet.cell(row=4, column=2).value)
print(c2_cell.row, c2_cell.column)


print(sheet['A5'].data_type)
print(sheet['B5'].data_type)


print(sheet['A5'].encoding)

print(sheet['D4'].parent)


print(dir(b2_cell))


cell_range = sheet['B2:C11']


for product, price in cell_range:
    print(f'Product: {product.value}    Price:{price.value}')


print(f'Sheet Dimentions: {sheet.dimensions}')
print(sheet.max_row, sheet.max_column)


for a, b, c, d, e in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, d.value, e.value)

for row in sheet.rows:
    for cell in row:
        print(f'{cell.value} ', end='')
    print('\n')

for row in sheet.values:
    print(row)

print('\n')
sheet['d2'] = 400

new_product = (11, 'Tablet', 12, 600, 12 * 600)  # this is a tuple
sheet.append(new_product)


for c, d, e in sheet['c2:e12']:
    e.value = c.value * d.value

wb.save('store.xlsx')
print('\n')
sheet['A1'] = 'Year'
sheet['B1'] = 'Sales'

sales = {2017: 700000, 2018: 800000, 2019: 900000}


for k, v in sales.items():
    sheet.append((k, v))

wb.save('sales.xlsx')

print('\n')

sheet = wb['Products']

for c, d, e in sheet['c2:e12']:
    e.value = f'={c.coordinate}*{d.coordinate}'
sheet = wb['Sales 2018']

sheet['B14'] = '=sum(B2:B13)'

wb.save('store.xlsx')

print(wb.sheetnames)

sheet = wb['Products']

print(dir(sheet))

#
sheet.title = 'Products for sale'



print(sheet.sheet_format)
print(sheet.sheet_properties)

wb.create_sheet('Turnover1', 0)

sheet1 = wb['Turnover1']

wb.remove(sheet1)

source = wb['Turnover']
destination = wb.copy_worksheet(source)
print(destination.title)

wb.save('store.xlsx')

from openpyxl.styles import *


from copy import copy

wb = openpyxl.load_workbook('store.xlsx')

sheet = wb['Products']

my_cell = sheet['B4']
# Challenges
sheet = wb.active
sheet.title = 'COMPANY SALES'

rows = (
    ('Year', 'Sales'),
    (2017, 150000),
    (2018, 180000),
    (2019, 210000),
    (2020, 125000),

)

for r in rows:
    sheet.append(r)


wb.save('sales.xlsx')

print(dir(openpyxl.styles))

font = Font(name='Tahoma', size=16, color=colors.RED, bold=True, italic=True, strike=False)
my_cell.font = font

fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
my_cell.fill = fill

double_border_green = Side(border_style='double', color=colors.GREEN)
thin_border_red = Side(border_style='thin', color='FF0000')
my_cell.border = Border(left=double_border_green, right=thin_border_red, top=double_border_green,
                        bottom=thin_border_red)

alignment = Alignment(horizontal='right', vertical='center')
my_cell.alignment = alignment


new_cell = sheet['B10']
new_font = copy(my_cell.font)
new_font.color = colors.GREEN
new_cell.font = new_font

wb.save('store.xlsx')


print('\n')
sheet = wb.active

items = list()
for row in sheet.values:
    items.append(row)

print(items)


vat = list()
for row in items[1:]:
    element = (row[0], row[1] * 0.15)
    vat.append(element)

print(vat)


wb.create_sheet('VAT')
sheet = wb['VAT']

sheet['A1'] = 'Year'
sheet['B1'] = 'VAT'

for row in vat:
    sheet.append(row)

wb = openpyxl.load_workbook('sales2.xlsx')

sheet = wb.active


cell = sheet['B6']
cell.value = '=sum(B2:B5)'


cell = sheet['C6']

cell.value = 'Total Sales'
font = Font(name='Tahoma', size=16, color=colors.RED, bold=True, italic=False, strike=False)
cell_b6 = sheet['B6']
cell_b6.font = font

cell_c6 = sheet['c6']
cell_c6.font = font

def csv2excel(filein, fileout, delim=','):
    import openpyxl, csv
    with open(filein, 'r') as f:
        reader = csv.reader(f, delimiter=delim)

        wb = openpyxl.Workbook()
        sheet = wb.active
        for row in reader:
            sheet.append(row)

        wb.save(fileout)

csv2excel('people3.csv', 'teachers.xlsx')

wb.save('sales3.xlsx')

def excel2csv(filein, fileout):
    import openpyxl, csv
    wb = openpyxl.load_workbook(filein)
    sheet = wb.active
    with open(fileout, 'w') as f:
        writer = csv.writer(f)
        for row in sheet.values:
            writer.writerow(row)

excel2csv('books.xlsx', 'booklist.csv')